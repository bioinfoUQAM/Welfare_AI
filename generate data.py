# -*- coding: utf-8 -*-
"""
Created on Thu Jun 25 20:04:34 2020

@author: Yesmin
"""


import pandas as pd
import matplotlib.pyplot as plt
from fitter import Fitter
import numpy as np
import scipy
from scipy.signal import medfilt
import random
from matplotlib.legend_handler import HandlerLine2D
from openpyxl import load_workbook
import pandas as pd
import os

fileNames_VideoNames = pd.read_excel(r"D:\BA_Yasmine_UQAM\Dictionary_Kinematics.xlsx", "Video File -> Excel Tab Names")
fileNames_VideoNames= fileNames_VideoNames.to_numpy()
side1Names = fileNames_VideoNames[0::2]
side2Names = fileNames_VideoNames[1::2]#start from the first element ald take every 2nd one
sheets=side1Names[:,0]
sheets2=side2Names[:,0]
df = [pd.DataFrame(pd.read_excel(r"D:\BA_Yasmine_UQAM\Plot\ScaledCoordinates_Post-Trial.xlsx",sheet).iloc[:, 0:34]) for x,sheet in enumerate(sheets)]  
data= df[0]['Z-E'].dropna()
columns = df[1].iloc[0, 2:34].index 
#generate synthetic sample by adding random 10%
generated_data=[]
for i, row in enumerate(data):
     bmin=int(min(-0.013*row, 0.013*row))
     bmax=-bmin
     generated_data.append(row + random.randint(bmin, bmax))
path = os.path.join(os.path.dirname(__file__), 'ScaledCoordinates_Post-Trial - Copie.xlsx')
# df1 = pd.DataFrame(generated_data)

# x2 = np.random.randn(100, 2)
# df2 = pd.DataFrame(x2)

# writer = pd.ExcelWriter(path, engine = 'xlsxwriter')
# df1.to_excel(writer, sheet_name = 'Z-E')
# df2.to_excel(writer, sheet_name = 'x2')
# writer.save()
# writer.close()

# writer = pd.ExcelWriter('ScaledCoordinates_Post-Trial - Copie.xlsx', engine='openpyxl')

# if os.path.exists('ScaledCoordinates_Post-Trial - Copie.xlsx'):
#     book = load_workbook('ScaledCoordinates_Post-Trial - Copie.xlsx')
#     writer.book = book
# df= pd.DataFrame(generated_data)
# df.to_excel(writer, sheet_name='new')
# writer.save()
# writer.close()

wb2 = load_workbook('ScaledCoordinates_Post-Trial - Copie.xlsx')
wb2.create_sheet('sid1')
wb2.save('ScaledCoordinates_Post-Trial - Copie.xlsx')


line1, = plt.plot(medfilt(generated_data), label='generated_data Z_E')

line2, = plt.plot(data, label='real_data Z_E')
plt.legend(handler_map={line1: HandlerLine2D(numpoints=4)})
plt.show()
ax=data.plot.hist(bins=15)